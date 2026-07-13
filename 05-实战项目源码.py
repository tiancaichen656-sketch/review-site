"""
============================================================
5 个接单级 Python 实战项目源码
测评导航 | tiancaichen656-sketch.github.io/review-site
============================================================
包含：
  项目1：豆瓣电影爬虫
  项目2：Excel 批量处理工具
  项目3：微信自动回复机器人
  项目4：PDF 报表生成器
  项目5：数据可视化大屏
============================================================
"""

# ============================================================
# 项目 1：豆瓣电影 Top250 爬虫
# 用途：爬取豆瓣电影 Top250，保存为 CSV
# 接单场景：帮客户爬取公开网站数据
# ============================================================

import requests
from bs4 import BeautifulSoup
import csv
import time

def crawl_douban_top250():
    """爬取豆瓣电影 Top250"""
    movies = []
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }

    for page in range(0, 250, 25):
        url = f'https://movie.douban.com/top250?start={page}'
        print(f'正在爬取第 {page//25 + 1} 页...')

        resp = requests.get(url, headers=headers)
        soup = BeautifulSoup(resp.text, 'html.parser')

        for item in soup.find_all('div', class_='item'):
            title = item.find('span', class_='title').text
            rating = item.find('span', class_='rating_num').text
            quote_elem = item.find('span', class_='inq')
            quote = quote_elem.text if quote_elem else ''
            info = item.find('div', class_='bd').p.text.strip()

            movies.append({
                'title': title,
                'rating': rating,
                'quote': quote,
                'info': info.replace('\n', ' ').replace('\r', ''),
            })

        time.sleep(1)  # 礼貌爬虫，避免被封 IP

    # 保存为 CSV
    with open('douban_top250.csv', 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['title', 'rating', 'quote', 'info'])
        writer.writeheader()
        writer.writerows(movies)

    print(f'✅ 爬取完成！共 {len(movies)} 部电影，已保存到 douban_top250.csv')
    return movies

# ============================================================
# 项目 2：Excel 批量处理工具
# 用途：批量处理多个 Excel 文件的合并、统计、格式化
# 接单场景：帮客户处理报表数据（最高频接单需求）
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def merge_excel_files(folder_path, output_file='merged.xlsx'):
    """合并文件夹内所有 Excel 文件"""
    all_data = []
    headers = None

    for filename in os.listdir(folder_path):
        if not filename.endswith('.xlsx') or filename.startswith('~'):
            continue

        filepath = os.path.join(folder_path, filename)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # 读取数据（跳过表头）
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row) if headers is None else headers
            else:
                all_data.append(list(row))

        print(f'✅ 已读取: {filename} ({len(list(ws.iter_rows())) - 1} 行)')

    # 写入新文件
    wb = openpyxl.Workbook()
    ws = wb.active

    # 写表头 + 样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据
    for i, row_data in enumerate(all_data, 2):
        for j, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = thin_border

    # 自动调整列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_file)
    print(f'✅ 合并完成！共 {len(all_data)} 行数据，保存到 {output_file}')
    return output_file

def generate_summary(input_file, output_file='summary.xlsx'):
    """对 Excel 数据生成汇总统计"""
    import pandas as pd
    df = pd.read_excel(input_file)

    # 数值列统计
    summary = df.describe()

    with pd.ExcelWriter(output_file) as writer:
        df.to_excel(writer, sheet_name='原始数据', index=False)
        summary.to_excel(writer, sheet_name='统计汇总')

    print(f'✅ 汇总完成！已保存到 {output_file}')
    return output_file

# ============================================================
# 项目 3：微信自动回复机器人
# 用途：监控微信消息并自动回复
# 接单场景：客服自动回复、群管理机器人
# 注意：需要安装 itchat 或使用 wxauto（Windows）
# ============================================================

try:
    import itchat
    from itchat.content import TEXT

    def wechat_auto_reply():
        """微信自动回复（需扫码登录）"""
        @itchat.msg_register(TEXT)
        def text_reply(msg):
            # 自定义关键词回复
            keywords = {
                '你好': '你好！我是自动回复机器人 🤖',
                '价格': 'Python 入门课完整测评报告 ¥9.9，含 6 个附件。链接：tiancaichen656-sketch.github.io/review-site',
                '课程': '推荐风变编程（¥1298，最全面）或夜曲编程（¥499，最友好）。详细对比戳链接~',
                'B站': 'B站免费课推荐黑马程序员系列，700+集系统全面。但要做好没人帮你 debug 的心理准备！',
            }

            for keyword, reply in keywords.items():
                if keyword in msg['Text']:
                    return reply

            return '收到！机器人正在学习中，稍后回复~ 🤖（测评报告：tiancaichen656-sketch.github.io/review-site）'

        print('🤖 微信自动回复机器人已启动，扫码登录...')
        itchat.auto_login(hotReload=True)
        itchat.run()

    # 使用方法：取消下面注释即可运行
    # wechat_auto_reply()

except ImportError:
    print('💡 微信机器人需要安装 itchat：pip install itchat')

# ============================================================
# 项目 4：PDF 报表生成器
# 用途：从 Excel 数据生成带图表的 PDF 报表
# 接单场景：帮客户做月报/季报自动化
# ============================================================

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
import matplotlib.pyplot as plt
import pandas as pd
import os

def generate_pdf_report(data_file, output_file='report.pdf', title='数据分析报告'):
    """从 Excel 生成 PDF 报表"""
    # 1. 读取数据
    df = pd.read_excel(data_file) if data_file.endswith('.xlsx') else pd.read_csv(data_file)

    # 2. 生成图表（保存为临时图片）
    chart_file = 'temp_chart.png'

    fig, axes = plt.subplots(1, 2, figsize=(10, 4))
    numeric_cols = df.select_dtypes(include='number').columns

    if len(numeric_cols) >= 1:
        df[numeric_cols[0]].hist(ax=axes[0], bins=20, color='#2563eb', edgecolor='white')
        axes[0].set_title(f'{numeric_cols[0]} 分布')

    if len(numeric_cols) >= 2:
        axes[1].scatter(df[numeric_cols[0]], df[numeric_cols[1]], alpha=0.6, color='#2563eb')
        axes[1].set_xlabel(numeric_cols[0])
        axes[1].set_ylabel(numeric_cols[1])
        axes[1].set_title(f'{numeric_cols[0]} vs {numeric_cols[1]}')

    plt.tight_layout()
    plt.savefig(chart_file, dpi=150)
    plt.close()

    # 3. 创建 PDF
    doc = SimpleDocTemplate(output_file, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # 标题
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 12))

    # 基本信息
    story.append(Paragraph(f'数据行数：{len(df)} | 数据列数：{len(df.columns)}', styles['Normal']))
    story.append(Spacer(1, 10))

    # 统计摘要
    story.append(Paragraph('<b>📊 统计摘要</b>', styles['Heading2']))
    desc = df.describe().round(2)
    table_data = [['指标'] + list(desc.columns)]
    for idx in desc.index:
        table_data.append([idx] + [str(desc.loc[idx, col]) for col in desc.columns])

    t = Table(table_data[:8])  # 只显示前 8 行
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    # 图表
    if os.path.exists(chart_file):
        story.append(Paragraph('<b>📈 数据可视化</b>', styles['Heading2']))
        img = Image(chart_file, width=450, height=180)
        story.append(img)

    # 底部
    story.append(Spacer(1, 30))
    story.append(Paragraph('报告由测评导航自动生成 | tiancaichen656-sketch.github.io/review-site',
                           ParagraphStyle('footer', fontSize=8, textColor=colors.grey)))

    doc.build(story)
    # 清理临时文件
    if os.path.exists(chart_file):
        os.remove(chart_file)

    print(f'✅ PDF 报表已生成：{output_file}')
    return output_file

# ============================================================
# 项目 5：数据可视化大屏
# 用途：生成一个交互式数据大屏 HTML 页面
# 接单场景：帮客户做数据展示/汇报大屏
# ============================================================

def generate_dashboard(data_file, output_file='dashboard.html'):
    """生成数据可视化大屏 HTML"""
    import pandas as pd
    import json

    df = pd.read_excel(data_file) if data_file.endswith('.xlsx') else pd.read_csv(data_file)
    numeric_cols = df.select_dtypes(include='number').columns.tolist()

    # 统计数据
    stats = {
        'total_rows': len(df),
        'total_cols': len(df.columns),
        'numeric_cols': numeric_cols,
        'summary': {col: {
            'mean': round(df[col].mean(), 2),
            'max': round(df[col].max(), 2),
            'min': round(df[col].min(), 2),
        } for col in numeric_cols[:5]},
    }

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>数据大屏 | 测评导航</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'PingFang SC','Microsoft YaHei',sans-serif; background: #0f172a; color: #e2e8f0; min-height: 100vh; padding: 20px; }}
        .header {{ text-align: center; padding: 20px 0; border-bottom: 2px solid #1e293b; margin-bottom: 20px; }}
        .header h1 {{ font-size: 28px; background: linear-gradient(135deg, #38bdf8, #818cf8); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }}
        .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 20px; }}
        .stat-card {{ background: #1e293b; border-radius: 12px; padding: 20px; text-align: center; border: 1px solid #334155; }}
        .stat-card .value {{ font-size: 36px; font-weight: 800; color: #38bdf8; }}
        .stat-card .label {{ font-size: 13px; color: #94a3b8; margin-top: 4px; }}
        .summary-table {{ width: 100%; border-collapse: collapse; margin-top: 16px; font-size: 14px; }}
        .summary-table th,.summary-table td {{ border: 1px solid #334155; padding: 10px 14px; text-align: center; }}
        .summary-table th {{ background: #1e293b; color: #38bdf8; }}
        .footer {{ text-align: center; color: #64748b; font-size: 12px; margin-top: 30px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 数据可视化大屏</h1>
        <p style="color:#94a3b8;margin-top:8px;">数据概览 · 统计摘要</p>
    </div>
    <div class="stats">
        <div class="stat-card"><div class="value">{stats['total_rows']}</div><div class="label">总行数</div></div>
        <div class="stat-card"><div class="value">{stats['total_cols']}</div><div class="label">总列数</div></div>
        <div class="stat-card"><div class="value">{len(numeric_cols)}</div><div class="label">数值列</div></div>
    </div>
    <h3 style="margin:20px 0 10px;">📈 统计摘要</h3>
    <table class="summary-table">
        <tr><th>列名</th><th>平均值</th><th>最大值</th><th>最小值</th></tr>
        {''.join(f'<tr><td>{col}</td><td>{s["mean"]}</td><td>{s["max"]}</td><td>{s["min"]}</td></tr>' for col, s in stats['summary'].items())}
    </table>
    <div class="footer">© 测评导航 | tiancaichen656-sketch.github.io/review-site</div>
</body>
</html>'''

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f'✅ 数据大屏已生成：{output_file}')
    return output_file

# ============================================================
# 快速测试
# ============================================================
if __name__ == '__main__':
    print('=' * 60)
    print('  5 个实战项目已就绪')
    print('  使用方法：取消对应函数的注释，填入你的数据文件路径')
    print('=' * 60)
    print()
    print('  项目1：crawl_douban_top250()          → 爬取豆瓣Top250')
    print('  项目2：merge_excel_files("./data/")   → 合并Excel')
    print('  项目3：wechat_auto_reply()            → 微信机器人')
    print('  项目4：generate_pdf_report("data.xlsx") → 生成PDF')
    print('  项目5：generate_dashboard("data.xlsx") → 数据大屏')
    print()
    print('  💡 接单提示：')
    print('  闲鱼搜索"Python爬虫""Excel处理"看同行定价')
    print('  客单价 ¥50-200/单，接5-10单就回本')
    print('  完整指南：tiancaichen656-sketch.github.io/review-site')
